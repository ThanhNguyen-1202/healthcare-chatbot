#!/usr/bin/env python3
"""Chạy benchmark Excel qua Healthcare Chatbot RAG++.

Script đọc sheet ``scoring_ready`` có 5 cột:
    question, ground_truth, contexts_ground_truth, answer, contexts_answer

Mỗi câu hỏi được gửi tới ``POST /predict/ragpp``. Kết quả được ghi:
    answer          -> JSON list Top-k tên bệnh
    contexts_answer -> JSON list các tài liệu RAG thực tế được truy xuất

Ví dụ:
    python scripts/run_ddxplus_benchmark.py \
        --input DDXPlus_Benchmark_1000_Cau_Giong_File_Mau.xlsx \
        --output ket_qua_benchmark_1000.xlsx
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional
from urllib.parse import urlsplit, urlunsplit

try:
    import httpx
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Thiếu thư viện httpx. Hãy chạy: python -m pip install httpx openpyxl"
    ) from exc

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Thiếu thư viện openpyxl. Hãy chạy: python -m pip install openpyxl"
    ) from exc


REQUIRED_COLUMNS = (
    "question",
    "ground_truth",
    "contexts_ground_truth",
    "answer",
    "contexts_answer",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Chạy file benchmark Excel và điền answer/contexts_answer.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--input",
        default="DDXPlus_Benchmark_1000_Cau_Giong_File_Mau.xlsx",
        help="File Excel đầu vào.",
    )
    parser.add_argument(
        "--output",
        default="ket_qua_DDXPlus_1000_da_test.xlsx",
        help="File Excel kết quả. Nếu đã tồn tại, script tiếp tục từ file này.",
    )
    parser.add_argument("--sheet", default="scoring_ready", help="Tên sheet chứa dữ liệu.")
    parser.add_argument(
        "--api-url",
        default="http://127.0.0.1:8000/predict/ragpp",
        help="Endpoint RAG++ của backend.",
    )
    parser.add_argument("--top-k", type=int, default=3, help="Số bệnh cần lấy.")
    parser.add_argument(
        "--delay",
        type=float,
        default=2.2,
        help="Số giây nghỉ sau mỗi request. 2.2 giây phù hợp giới hạn 30 request/phút.",
    )
    parser.add_argument("--timeout", type=float, default=180.0, help="Timeout mỗi request.")
    parser.add_argument("--max-retries", type=int, default=5, help="Số lần thử lại khi lỗi.")
    parser.add_argument(
        "--save-every",
        type=int,
        default=5,
        help="Lưu checkpoint sau mỗi N dòng thành công hoặc lỗi.",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="Dòng Excel bắt đầu chạy, tính cả dòng tiêu đề ở dòng 1.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Chỉ chạy tối đa N dòng; 0 nghĩa là chạy hết.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Chạy lại cả những dòng đã có answer và contexts_answer.",
    )
    parser.add_argument(
        "--fresh",
        action="store_true",
        help="Không tiếp tục file output cũ; tạo lại từ file input.",
    )
    parser.add_argument(
        "--stop-on-error",
        action="store_true",
        help="Dừng ngay khi một dòng lỗi thay vì ghi log và chạy tiếp.",
    )
    return parser.parse_args()


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def build_health_url(api_url: str) -> str:
    parts = urlsplit(api_url)
    return urlunsplit((parts.scheme, parts.netloc, "/health", "", ""))


def unwrap_api_response(body: Dict[str, Any]) -> Dict[str, Any]:
    data = body.get("data")
    if isinstance(data, dict):
        return data
    return body


def clean_text(value: Any, max_length: int = 12000) -> str:
    text = str(value or "").replace("\x00", "").strip()
    if len(text) > max_length:
        return text[: max_length - 3] + "..."
    return text


def disease_names(payload: Dict[str, Any], top_k: int) -> List[str]:
    candidates = payload.get("top_diseases")
    if not isinstance(candidates, list):
        candidates = (payload.get("prediction") or {}).get("possible_diseases", [])

    names: List[str] = []
    for item in candidates or []:
        if isinstance(item, dict):
            name = item.get("name") or item.get("label") or item.get("canonical_name")
        else:
            name = item
        name = clean_text(name, 250)
        if name and name not in names:
            names.append(name)
        if len(names) >= top_k:
            break
    return names


def format_context(doc: Any) -> str:
    if not isinstance(doc, dict):
        return clean_text(doc)

    source = clean_text(doc.get("source"), 300)
    title = clean_text(doc.get("title"), 500)
    diseases = doc.get("disease_names") or []
    if isinstance(diseases, (list, tuple)):
        disease_text = ", ".join(clean_text(x, 200) for x in diseases if x)
    else:
        disease_text = clean_text(diseases, 500)

    content = clean_text(
        doc.get("content")
        or doc.get("text")
        or doc.get("summary")
        or doc.get("snippet"),
        8000,
    )
    score = doc.get("retrieval_score")

    parts: List[str] = []
    if source:
        parts.append(f"Nguồn: {source}")
    if title:
        parts.append(f"Tiêu đề: {title}")
    if disease_text:
        parts.append(f"Bệnh liên quan: {disease_text}")
    if score is not None:
        parts.append(f"Điểm truy xuất: {score}")
    if content:
        parts.append(content)

    # Giới hạn an toàn cho một ô Excel (tối đa 32.767 ký tự).
    return clean_text("\n".join(parts), 12000)


def rag_contexts(payload: Dict[str, Any]) -> List[str]:
    ragpp = payload.get("ragpp") or {}
    docs = ragpp.get("evidence") or ragpp.get("matched_documents") or []
    contexts = [format_context(doc) for doc in docs]
    return [text for text in contexts if text]


def atomic_save(workbook: Any, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_name(output_path.stem + ".tmp.xlsx")
    workbook.save(temp_path)
    os.replace(temp_path, output_path)


def append_error(error_path: Path, record: Dict[str, Any]) -> None:
    with error_path.open("a", encoding="utf-8") as stream:
        stream.write(json.dumps(record, ensure_ascii=False) + "\n")


def request_prediction(
    client: httpx.Client,
    api_url: str,
    question: str,
    top_k: int,
    max_retries: int,
) -> Dict[str, Any]:
    last_error: Optional[Exception] = None

    for attempt in range(1, max_retries + 1):
        try:
            response = client.post(
                api_url,
                json={"message": question, "top_k": top_k},
                headers={"Accept": "application/json"},
            )

            if response.status_code == 429:
                retry_after = response.headers.get("Retry-After")
                try:
                    wait_seconds = max(float(retry_after or 0), 60.0)
                except ValueError:
                    wait_seconds = 60.0
                print(
                    f"  API trả 429. Chờ {wait_seconds:.0f} giây rồi thử lại "
                    f"({attempt}/{max_retries})...",
                    flush=True,
                )
                time.sleep(wait_seconds)
                continue

            if response.status_code >= 500:
                raise RuntimeError(
                    f"HTTP {response.status_code}: {clean_text(response.text, 1000)}"
                )

            response.raise_for_status()
            body = response.json()
            if body.get("success") is False:
                raise RuntimeError(
                    f"API success=false: {clean_text(body.get('errors') or body.get('message'), 1500)}"
                )
            return unwrap_api_response(body)

        except (httpx.HTTPError, ValueError, RuntimeError) as exc:
            last_error = exc
            if attempt >= max_retries:
                break
            wait_seconds = min(5 * (2 ** (attempt - 1)), 60)
            print(
                f"  Lỗi request: {exc}. Thử lại sau {wait_seconds}s "
                f"({attempt}/{max_retries})...",
                flush=True,
            )
            time.sleep(wait_seconds)

    raise RuntimeError(f"Không gọi được API sau {max_retries} lần: {last_error}")


def main() -> int:
    args = parse_args()
    if args.top_k < 1 or args.top_k > 5:
        raise SystemExit("--top-k phải nằm trong khoảng 1 đến 5.")
    if args.start_row < 2:
        raise SystemExit("--start-row phải từ 2 trở lên vì dòng 1 là tiêu đề.")

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    error_path = output_path.with_suffix(".errors.jsonl")

    if not input_path.exists():
        raise SystemExit(f"Không tìm thấy file đầu vào: {input_path}")

    source_path = input_path
    if output_path.exists() and not args.fresh:
        source_path = output_path
        print(f"Tiếp tục từ checkpoint: {output_path}")
    else:
        print(f"Tạo kết quả mới từ: {input_path}")

    try:
        workbook = load_workbook(source_path)
    except PermissionError as exc:
        raise SystemExit(
            "Không mở được Excel. Hãy đóng file .xlsx đang mở rồi chạy lại."
        ) from exc

    if args.sheet not in workbook.sheetnames:
        raise SystemExit(
            f"Không có sheet '{args.sheet}'. Các sheet hiện có: {workbook.sheetnames}"
        )
    sheet = workbook[args.sheet]

    header_to_col: Dict[str, int] = {}
    for col in range(1, sheet.max_column + 1):
        header = normalize_header(sheet.cell(1, col).value)
        if header:
            header_to_col[header] = col

    missing_headers = [name for name in REQUIRED_COLUMNS if name not in header_to_col]
    if missing_headers:
        raise SystemExit(
            "File Excel thiếu cột bắt buộc: " + ", ".join(missing_headers)
        )

    question_col = header_to_col["question"]
    answer_col = header_to_col["answer"]
    contexts_col = header_to_col["contexts_answer"]

    candidate_rows: List[int] = []
    for row in range(args.start_row, sheet.max_row + 1):
        question = clean_text(sheet.cell(row, question_col).value, 30000)
        if not question:
            continue
        current_answer = clean_text(sheet.cell(row, answer_col).value, 100)
        current_contexts = clean_text(sheet.cell(row, contexts_col).value, 100)
        if not args.overwrite and current_answer and current_contexts:
            continue
        candidate_rows.append(row)

    if args.limit > 0:
        candidate_rows = candidate_rows[: args.limit]

    if not candidate_rows:
        print("Không còn dòng nào cần chạy.")
        return 0

    print(f"Số dòng cần chạy: {len(candidate_rows)}")
    print(f"API: {args.api_url}")
    print(f"Output: {output_path}")
    print("Lưu ý: không mở file output bằng Excel trong lúc script đang chạy.\n")

    timeout = httpx.Timeout(args.timeout, connect=min(args.timeout, 30.0))
    completed = 0
    failed = 0
    started_at = time.time()

    with httpx.Client(timeout=timeout, follow_redirects=True) as client:
        health_url = build_health_url(args.api_url)
        try:
            health = client.get(health_url)
            health.raise_for_status()
            print(f"Backend hoạt động: {health_url}\n")
        except Exception as exc:
            print(
                f"CẢNH BÁO: chưa kiểm tra được {health_url}: {exc}\n"
                "Script vẫn tiếp tục thử gọi endpoint dự đoán.\n",
                flush=True,
            )

        for position, row in enumerate(candidate_rows, start=1):
            question = clean_text(sheet.cell(row, question_col).value, 30000)
            preview = question.replace("\n", " ")[:90]
            print(
                f"[{position}/{len(candidate_rows)}] Dòng {row}: {preview}",
                flush=True,
            )

            try:
                payload = request_prediction(
                    client=client,
                    api_url=args.api_url,
                    question=question,
                    top_k=args.top_k,
                    max_retries=args.max_retries,
                )
                names = disease_names(payload, args.top_k)
                contexts = rag_contexts(payload)

                if not names:
                    raise RuntimeError("API không trả về top_diseases/possible_diseases.")
                if not contexts:
                    raise RuntimeError("API không trả về ragpp.evidence.")

                sheet.cell(row, answer_col).value = json.dumps(
                    names, ensure_ascii=False
                )
                sheet.cell(row, contexts_col).value = json.dumps(
                    contexts, ensure_ascii=False
                )
                completed += 1
                print(
                    f"  OK answer={names}; contexts={len(contexts)}",
                    flush=True,
                )

            except Exception as exc:
                failed += 1
                record = {
                    "time": datetime.now().isoformat(timespec="seconds"),
                    "row": row,
                    "question": question,
                    "error": str(exc),
                }
                append_error(error_path, record)
                print(f"  LỖI: {exc}", file=sys.stderr, flush=True)
                if args.stop_on_error:
                    atomic_save(workbook, output_path)
                    raise

            processed = completed + failed
            if processed % max(args.save_every, 1) == 0:
                atomic_save(workbook, output_path)
                print(f"  Đã lưu checkpoint: {output_path.name}", flush=True)

            if args.delay > 0 and position < len(candidate_rows):
                time.sleep(args.delay)

    atomic_save(workbook, output_path)
    elapsed = time.time() - started_at
    print("\n========== HOÀN TẤT ==========")
    print(f"Thành công : {completed}")
    print(f"Lỗi       : {failed}")
    print(f"Thời gian : {elapsed / 60:.1f} phút")
    print(f"Kết quả   : {output_path}")
    if failed:
        print(f"Log lỗi   : {error_path}")
        print("Chạy lại cùng lệnh để tự động tiếp tục các dòng còn thiếu.")
    return 0 if failed == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
